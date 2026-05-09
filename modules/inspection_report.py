from utils.gemini_client import call_gemini

# Regulatory references by inspection type for grounding the findings
_REG_REFS = {
    "GMP Inspection": (
        "Schedule M (GMP for premises and materials), Drugs & Cosmetics Act 1940, "
        "WHO GMP TRS 986, ICH Q7, Q10. Reference sections as Schedule M Rule numbers."
    ),
    "GCP Inspection": (
        "ICH E6(R2) Good Clinical Practice, CDSCO GCP Guidelines 2001, "
        "Schedule Y (clinical trial requirements). Reference ICH E6 sections."
    ),
    "GDP Inspection": (
        "WHO GDP Guidelines (TRS 957), CDSCO distribution guidelines. "
        "Reference WHO annex numbers."
    ),
    "Pharmacovigilance Audit": (
        "CDSCO Pharmacovigilance Programme of India (PvPI), ICH E2A, E2C, E2E, "
        "Schedule Y adverse event reporting requirements."
    ),
    "Clinical Trial Site Audit": (
        "CDSCO GCP guidelines, ICH E6(R2), Schedule Y, New Drugs and Clinical Trials Rules 2019. "
        "Reference NDCT Rules clause numbers."
    ),
}

# Inspection area codes per CDSCO inspection SOP QMS-INS-004
_INSPECTION_AREAS = {
    "GMP Inspection": [
        "Pharmaceutical Quality System / QMS",
        "Personnel",
        "Premises, Equipment & Utilities",
        "Documentation",
        "Production & Process Monitoring",
        "Quality Control",
    ],
    "GCP Inspection": [
        "Protocol & Amendments",
        "Informed Consent",
        "Subject Records & Case Report Forms",
        "Investigational Product Management",
        "Ethics Committee Oversight",
        "Investigator Qualifications & Training",
        "Safety Reporting",
    ],
    "GDP Inspection": [
        "Storage Conditions & Temperature Control",
        "Documentation & Traceability",
        "Transportation & Cold Chain",
        "Personnel & Training",
        "Returns & Recalls",
    ],
    "Pharmacovigilance Audit": [
        "SAE / ICSR Reporting Procedures",
        "Signal Detection & Management",
        "PSUR / DSUR Preparation",
        "Qualified Person for Pharmacovigilance (QPPV)",
        "Training & SOPs",
    ],
    "Clinical Trial Site Audit": [
        "Protocol Compliance & Eligibility",
        "Informed Consent Process",
        "Source Document Verification",
        "Investigational Product Accountability",
        "SAE Reporting",
        "Ethics Committee Communication",
    ],
}

PROMPT = """You are a CDSCO-qualified pharmaceutical inspector writing a formal inspection report for official submission.
Convert the raw observations below into a structured report following CDSCO's standard format.

Inspection Type: {inspection_type}
Applicable Regulations: {reg_refs}
Standard Inspection Areas for this type: {inspection_areas}

=== SCOPE OF INSPECTION ===
Extract the following scope details from the observations:
- systems_covered: which of the standard inspection areas were reviewed
- product_types: specific product categories / dosage forms inspected
- manufacturing_lines: specific lines, rooms, or units visited (or trial sites for GCP)
- inspection_dates: date or date range of the inspection
- summary: one sentence describing what was in scope

=== FINDING CLASSIFICATION DEFINITIONS ===
Critical    — deficiency likely to result in a serious risk to patient safety or public health; falsification of data; systemic failure that invalidates product batches or trial data; requires immediate action
Major       — substantial deficiency that may compromise product quality, patient safety, or data integrity, but does not pose an immediate risk; could become critical if uncorrected
Minor       — deficiency that does not compromise quality or safety but indicates imperfect adherence to GxP; requires correction
Observation — area noted for potential improvement; not a regulatory violation; no formal response required but good practice to address

=== FINDING ID FORMAT ===
Use a code that indicates the area, e.g.:
  OBS-ICF-001  (Informed Consent Form)
  OBS-CRF-001  (Case Report Form / data)
  OBS-DEV-001  (Protocol deviation / eligibility)
  OBS-DRUG-001 (Investigational product / dispensing)
  OBS-EC-001   (Ethics Committee)
  OBS-TRN-001  (Training / qualification)
  OBS-SOP-001  (SOPs / documentation)
  OBS-TEMP-001 (Temperature / storage — for GMP/GDP)
  OBS-LABEL-001(Labelling)
  OBS-QC-001   (Quality control / testing)
  OBS-QMS-001  (Quality Management System)
  OBS-PROD-001 (Production / manufacturing process)
  OBS-PERS-001 (Personnel / qualification)
  OBS-PREM-001 (Premises / equipment)

=== DESCRIPTION QUALITY STANDARD ===
Each finding description must:
  1. Name the specific subject/batch/document/person involved (anonymise patient IDs to IN-XXX-XXXX format if present)
  2. State the observed fact precisely (dates, values, counts if available)
  3. State what was required (reference the specific rule, SOP, or protocol requirement)
  4. State why this matters (patient safety risk or data integrity consequence)

=== PROPOSED CAPA QUALITY STANDARD ===
Each CAPA must specify:
  - Immediate corrective action (what must be done now)
  - Root cause investigation required (yes/no, and what to investigate)
  - Preventive measure (systemic change to prevent recurrence)
  - Suggested timeline (e.g. "within 7 days", "within 30 days", "at next training cycle")

=== COMPLIANCE DETERMINATION ===
  Compliant               — no Critical or Major findings; Minor findings noted for correction
  Conditionally Compliant — Major findings present; CAPA accepted within specified timeframe
  Non-Compliant           — Critical findings present; or persistent non-compliance; or data integrity concerns

=== FOLLOW-UP TIMELINE ===
  Critical findings: CAPA response within 15 working days
  Major findings:    CAPA response within 30 days
  Minor only:        Response within 60 days or at next scheduled inspection

Return ONLY valid JSON, no markdown, no commentary:
{{
  "report_header": {{
    "inspection_type": "{inspection_type}",
    "facility_name": "full facility name",
    "facility_address": "city/address as stated",
    "inspection_date": "date or date range",
    "inspectors": ["Inspector Name — designation/organisation"],
    "report_date": "date report finalised"
  }},
  "scope": {{
    "systems_covered": ["system 1", "system 2"],
    "product_types": ["product/dosage form 1"],
    "manufacturing_lines": ["line or unit 1"],
    "inspection_dates": "date or date range",
    "summary": "one sentence scope description"
  }},
  "executive_summary": "3-4 sentences: what was inspected, key deficiency themes, overall compliance conclusion, and immediate implications",
  "findings": [
    {{
      "finding_id": "OBS-XXX-001",
      "category": "Critical"|"Major"|"Minor"|"Observation",
      "area": "inspection area",
      "description": "detailed observation",
      "regulatory_reference": "rule violated",
      "risk_level": "High"|"Medium"|"Low",
      "proposed_capa": "corrective action plan"
    }}
  ],
  "gmp_compliance": "Compliant"|"Conditionally Compliant"|"Non-Compliant",
  "critical_findings_count": integer,
  "major_findings_count": integer,
  "minor_findings_count": integer,
  "risk_assessment": [
    {{
      "risk_id": "RSK-001",
      "risk_title": "short title",
      "event_statement": "Cause -> Event -> Impact",
      "likelihood": 1-5,
      "severity": 1-5,
      "detectability": 1-5,
      "risk_priority_score": integer (likelihood * severity * detectability),
      "controls": "existing controls",
      "action_plan": "recommended recovery actions"
    }}
  ],
  "overall_assessment": "2-3 sentences on systemic compliance posture, risk to subjects/data, and any patterns across findings",
  "recommendations": ["specific actionable recommendation — name the responsible party and timeline"],
  "follow_up_required": true|false,
  "follow_up_timeline": "specific timeline per classification rules above"
}}

Raw Observations:
\"\"\"{text}\"\"\"
"""


def generate_inspection_report(text: str, report_type: str, client, model_name: str) -> dict:
    reg_refs = _REG_REFS.get(report_type, _REG_REFS["GMP Inspection"])
    areas = ", ".join(_INSPECTION_AREAS.get(report_type, []))
    prompt = PROMPT.format(
        inspection_type=report_type,
        reg_refs=reg_refs,
        inspection_areas=areas,
        text=text[:150000],
    )
    result = call_gemini(client, model_name, prompt, fallback={"error": "Report generation failed"})
    # Populate count fields if Gemini didn't fill them
    if "findings" in result and isinstance(result["findings"], list):
        cats = [f.get("category", "") for f in result["findings"]]
        result.setdefault("critical_findings_count", cats.count("Critical"))
        result.setdefault("major_findings_count", cats.count("Major"))
        result.setdefault("minor_findings_count", cats.count("Minor"))
    return result


def format_report_as_text(report: dict) -> str:
    h = report.get("report_header", {})
    sc = report.get("scope", {})
    lines = [
        "=" * 70,
        "CENTRAL DRUGS STANDARD CONTROL ORGANISATION (CDSCO)",
        "INSPECTION REPORT",
        "=" * 70,
        f"Inspection Type : {h.get('inspection_type','N/A')}",
        f"Facility Name   : {h.get('facility_name','N/A')}",
        f"Address         : {h.get('facility_address','N/A')}",
        f"Inspection Date : {h.get('inspection_date','N/A')}",
        f"Report Date     : {h.get('report_date','N/A')}",
        f"Inspectors      : {', '.join(h.get('inspectors',[])) or 'N/A'}",
    ]

    if sc:
        lines += [
            "",
            "SCOPE OF INSPECTION",
            "-" * 40,
            f"  {sc.get('summary', '')}",
        ]
        if sc.get("systems_covered"):
            lines.append(f"  Systems Covered : {', '.join(sc['systems_covered'])}")
        if sc.get("product_types"):
            lines.append(f"  Products        : {', '.join(sc['product_types'])}")
        if sc.get("manufacturing_lines"):
            lines.append(f"  Lines/Units     : {', '.join(sc['manufacturing_lines'])}")

    lines += [
        "",
        "EXECUTIVE SUMMARY",
        "-" * 40,
        report.get("executive_summary",""),
        "",
        f"GMP Compliance: {report.get('gmp_compliance','N/A')}",
        f"Findings: {report.get('critical_findings_count',0)} Critical | "
        f"{report.get('major_findings_count',0)} Major | "
        f"{report.get('minor_findings_count',0)} Minor",
        "",
        "DETAILED FINDINGS",
        "-" * 40,
    ]
    for f in report.get("findings", []):
        lines += [
            f"\n[{f.get('finding_id','')}] {f.get('category','').upper()} — Risk: {f.get('risk_level','')}",
            f"  Area           : {f.get('area','')}",
            f"  Description    : {f.get('description','')}",
            f"  Regulatory Ref : {f.get('regulatory_reference','')}",
        ]
        if f.get("corrective_action_required"):
            lines.append(f"  CAPA           : {f.get('proposed_capa','')}")
    lines += [
        "", "OVERALL ASSESSMENT", "-" * 40,
        report.get("overall_assessment",""),
        "", "RECOMMENDATIONS", "-" * 40,
    ]
    for i, r in enumerate(report.get("recommendations",[]), 1):
        lines.append(f"  {i}. {r}")
    fu = report.get("follow_up_required", False)
    lines.append(f"\nFollow-up: {'Yes — ' + report.get('follow_up_timeline','') if fu else 'No'}")
    lines.append("=" * 70)
    return "\n".join(lines)


def format_report_as_xlsx(report: dict) -> bytes:
    import io
    import pandas as pd
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Summary Sheet
        h = report.get("report_header", {})
        sc = report.get("scope", {})
        summary_data = [
            ["CENTRAL DRUGS STANDARD CONTROL ORGANISATION (CDSCO)", ""],
            ["INSPECTION REPORT SUMMARY", ""],
            ["", ""],
            ["Inspection Type", h.get("inspection_type", "N/A")],
            ["Facility Name", h.get("facility_name", "N/A")],
            ["Address", h.get("facility_address", "N/A")],
            ["Inspection Date", h.get("inspection_date", "N/A")],
            ["Report Date", h.get("report_date", "N/A")],
            ["Inspectors", ", ".join(h.get("inspectors", []))],
            ["", ""],
            ["EXECUTIVE SUMMARY", ""],
            [report.get("executive_summary", ""), ""],
            ["", ""],
            ["COMPLIANCE STATUS", report.get("gmp_compliance", "N/A")],
            ["Critical Findings", report.get("critical_findings_count", 0)],
            ["Major Findings", report.get("major_findings_count", 0)],
            ["Minor Findings", report.get("minor_findings_count", 0)],
        ]
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Report Summary", index=False, header=False)

        # 2. Detailed Findings Sheet
        findings = report.get("findings", [])
        if findings:
            df_findings = pd.DataFrame(findings)
            # Rename columns for better readability
            col_map = {
                "finding_id": "ID", "category": "Category", "area": "Inspection Area",
                "description": "Observation Description", "regulatory_reference": "Regulatory Reference",
                "risk_level": "Risk Level", "proposed_capa": "Proposed CAPA"
            }
            df_findings = df_findings.rename(columns=col_map)[[c for c in col_map.values() if c in df_findings.rename(columns=col_map).columns]]
            df_findings.to_excel(writer, sheet_name="Detailed Findings", index=False)

        # 3. Risk Register Sheet (matching the Denso style)
        risk_data = report.get("risk_assessment", [])
        if risk_data:
            df_risk = pd.DataFrame(risk_data)
            col_map_risk = {
                "risk_id": "Risk ID", "risk_title": "Risk Title", 
                "event_statement": "Risk Event Statement (Cause → Event → Impact)",
                "likelihood": "Likelihood (1-5)", "severity": "Severity (1-5)", 
                "detectability": "Detectability (1-5)", "risk_priority_score": "Risk Priority Score",
                "controls": "Controls (Prevent/Detect)", "action_plan": "Action Plan (Correct/Recover)"
            }
            df_risk = df_risk.rename(columns=col_map_risk)
            df_risk.to_excel(writer, sheet_name="Risk Register", index=False)

        # Basic styling via openpyxl
        workbook = writer.book
        # Style Summary
        ws_sum = workbook["Report Summary"]
        ws_sum.column_dimensions['A'].width = 30
        ws_sum.column_dimensions['B'].width = 80
        for cell in ws_sum[1]: cell.font = Font(bold=True, size=14)
        for cell in ws_sum[2]: cell.font = Font(bold=True, size=12)

        # Style Findings
        if "Detailed Findings" in workbook.sheetnames:
            ws_find = workbook["Detailed Findings"]
            for cell in ws_find[1]: 
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F8EF7", end_color="4F8EF7", fill_type="solid")
            for col in ws_find.columns:
                ws_find.column_dimensions[col[0].column_letter].width = 25

        # Style Risk Register
        if "Risk Register" in workbook.sheetnames:
            ws_risk = workbook["Risk Register"]
            for cell in ws_risk[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="161B2E", end_color="161B2E", fill_type="solid")
            for col in ws_risk.columns:
                ws_risk.column_dimensions[col[0].column_letter].width = 20

    return output.getvalue()

